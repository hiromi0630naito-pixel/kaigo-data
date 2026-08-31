#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
架電チャネルの損益計画モデルを作る。

数量（リスト件数・ファネル）は検証済みの値と評価基準の目標値を使い、
金額（単価・人件費・固定費）は空欄で出す。埋めると月次P&Lが自動計算される。

使い方:
    python3 scripts/build_pl_model.py
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "pl_model.xlsx"
MONTHS = 12

HEAD = PatternFill("solid", fgColor="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")    # 入力してほしいセル
FIXED_FILL = PatternFill("solid", fgColor="E2EFDA")    # 確定値
WHITE_BOLD = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 入力シートのセル位置。他シートから参照するので定数にしておく。
I = {
    "リスト件数": "B4", "稼働人数": "B5", "1日架電数": "B6", "稼働日数": "B7",
    "無効率": "B8", "アポ率": "B9", "商談率": "B10", "受注率": "B11",
    "初回単価": "B15", "月額": "B16", "継続月数": "B17", "変動費率": "B18",
    "人件費": "B19", "ツール費": "B20", "その他固定費": "B21",
    "枯渇後固定費": "B22",
}


def ref(key):
    return "入力!$" + I[key].replace("B", "B$")


def _head(ws, row, text, span=6):
    c = ws.cell(row, 1, text)
    c.font = WHITE_BOLD
    c.fill = HEAD
    for i in range(2, span + 1):
        ws.cell(row, i).fill = HEAD


def build_input(ws):
    ws["A1"] = "入力"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "黄色のセルを埋めてください。緑は検証済みの確定値です。"
    ws["A2"].font = Font(italic=True, color="808080")

    _head(ws, 3, "数量", 4)
    rows = [
        ("対象リスト件数", 1308, "件", "確定。tokyo_minkan_1jigyosho.csv の行数", "fix"),
        ("稼働人数", 3, "人", "しょうちゃん・さくや・廣澤", "fix"),
        ("1人1日あたり架電数", 50, "件", "実測が出たら置き換える", "in"),
        ("月間稼働日数", 20, "日", "", "in"),
        ("無効リード率", 0.03, "率", "番号違い・対象外。評価基準の目標値（仮）", "in"),
        ("アポ率(有効架電比)", 0.03, "率", "評価基準の目標値（仮）", "in"),
        ("商談実施率", 0.80, "率", "評価基準の目標値（仮）", "in"),
        ("受注率(商談比)", 0.20, "率", "評価基準の目標値（仮）", "in"),
    ]
    for i, (k, v, u, note, kind) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, k).font = Font(bold=True)
        c = ws.cell(r, 2, v)
        c.fill = FIXED_FILL if kind == "fix" else INPUT_FILL
        c.border = BOX
        c.number_format = "0.0%" if u == "率" else "#,##0"
        ws.cell(r, 3, u)
        ws.cell(r, 4, note).font = Font(size=10, color="808080")

    _head(ws, 13, "金額（未入力）", 4)
    money = [
        ("初回単価", "円", "1件受注したときの初回売上"),
        ("月額継続課金", "円/月", "無ければ0"),
        ("平均継続月数", "ヶ月", "月額課金がある場合。無ければ0"),
        ("変動費率", "率", "原価・外注費の売上に対する比率"),
        ("人件費(1人あたり月)", "円", "3人分は自動で掛ける"),
        ("通信・ツール費(月)", "円", "電話・CRM等"),
        ("その他固定費(月)", "円", "家賃・雑費など"),
    ]
    for i, (k, u, note) in enumerate(money):
        r = 15 + i
        ws.cell(r, 1, k).font = Font(bold=True)
        c = ws.cell(r, 2)
        c.fill = INPUT_FILL
        c.border = BOX
        c.number_format = "0.0%" if u == "率" else "#,##0"
        ws.cell(r, 3, u)
        ws.cell(r, 4, note).font = Font(size=10, color="808080")

    ws.cell(22, 1, "リスト枯渇後の固定費").font = Font(bold=True)
    sw = ws.cell(22, 2, "計上する")
    sw.fill = INPUT_FILL
    sw.border = BOX
    dv = DataValidation(type="list", formula1='"計上する,計上しない"',
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("B22")
    ws.cell(22, 3, "選択")
    ws.cell(22, 4, "人員を維持するなら「計上する」。"
                   "架電が止まっても人件費は消えない").font = \
        Font(size=10, color="808080")

    _head(ws, 23, "自動計算", 4)
    calc = [
        ("月間架電キャパ", f"={ref('稼働人数')}*{ref('1日架電数')}*{ref('稼働日数')}",
         "本/月", "3人が1ヶ月にかけられる本数"),
        ("架電→受注の転換率",
         f"=(1-{ref('無効率')})*{ref('アポ率')}*{ref('商談率')}*{ref('受注率')}",
         "率", "架電1本が受注になる確率"),
        ("1受注あたり必要架電数", "=IFERROR(1/B24,\"\")", "本", ""),
        ("月間固定費計",
         f"={ref('人件費')}*{ref('稼働人数')}+{ref('ツール費')}"
         f"+{ref('その他固定費')}", "円/月", "金額を入れると出ます"),
    ]
    for i, (k, f, u, note) in enumerate(calc):
        r = 24 + i
        ws.cell(r, 1, k).font = Font(bold=True)
        c = ws.cell(r, 2, f)
        c.border = BOX
        c.number_format = "0.000%" if u == "率" else "#,##0"
        ws.cell(r, 3, u)
        ws.cell(r, 4, note).font = Font(size=10, color="808080")

    for col, w in zip("ABCD", (24, 16, 8, 46)):
        ws.column_dimensions[col].width = w


PL_COLS = [
    ("月", None),
    ("期首残リスト", "=IF(A{r}=1,{list},H{p})"),
    ("架電数", "=MIN({cap},B{r})"),
    ("有効架電", "=C{r}*(1-{muko})"),
    ("アポ", "=D{r}*{apo}"),
    ("商談", "=E{r}*{sho}"),
    ("受注", "=F{r}*{ju}"),
    ("期末残リスト", "=B{r}-C{r}"),
    ("新規売上", "=G{r}*{tanka}"),
    ("継続売上", "=SUMIFS($G$4:$G${last},$A$4:$A${last},\">=\"&A{r}-{keizoku}+1,"
                "$A$4:$A${last},\"<=\"&A{r})*{getsugaku}"),
    ("売上計", "=I{r}+J{r}"),
    ("変動費", "=K{r}*{hendo}"),
    ("粗利", "=K{r}-L{r}"),
    ("固定費", "=IF(OR({switch}=\"計上する\",C{r}>0),{kotei},0)"),
    ("営業利益", "=M{r}-N{r}"),
    ("累計営業利益", "=IF(A{r}=1,O{r},P{p}+O{r})"),
]


def build_pl(ws):
    ws["A1"] = "月次P&L"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("リストが尽きると架電数が自動で頭打ちになります。"
                "枯渇後も人員を維持する前提なら固定費は計上され続けます"
                "（入力シートで切り替え）。")
    ws["A2"].font = Font(italic=True, color="808080")

    for i, (name, _) in enumerate(PL_COLS):
        c = ws.cell(3, i + 1, name)
        c.font = WHITE_BOLD
        c.fill = HEAD
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    last = 3 + MONTHS
    subs = {
        "list": ref("リスト件数"), "cap": "入力!$B$24", "muko": ref("無効率"),
        "apo": ref("アポ率"), "sho": ref("商談率"), "ju": ref("受注率"),
        "tanka": ref("初回単価"), "getsugaku": ref("月額"),
        "keizoku": ref("継続月数"), "hendo": ref("変動費率"),
        "kotei": "入力!$B$27", "switch": ref("枯渇後固定費"), "last": last,
    }
    for m in range(1, MONTHS + 1):
        r = 3 + m
        ws.cell(r, 1, m)
        for i, (_, tpl) in enumerate(PL_COLS[1:], start=2):
            f = tpl.format(r=r, p=r - 1, **subs)
            c = ws.cell(r, i, f)
            c.border = BOX
            c.number_format = "#,##0"
    for i in range(1, len(PL_COLS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13 if i > 1 else 6
    ws.freeze_panes = "B4"


def build_breakeven(ws):
    ws["A1"] = "損益分岐"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("上の表は単価を入れなくても読めます。"
                "月次固定費と月あたり受注件数から、必要な単価が出ます。")
    ws["A2"].font = Font(italic=True, color="808080")

    _head(ws, 4, "必要単価の逆算（粗利率は入力シートの変動費率から算出）", 7)
    ws.cell(5, 1, "月次固定費 \\ 月間受注件数").font = Font(bold=True)
    counts = [1, 2, 4, 6, 10, 15]
    for j, n in enumerate(counts):
        c = ws.cell(5, 2 + j, f"{n}件")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    for i, fc in enumerate([300000, 600000, 1000000, 1500000, 2000000]):
        r = 6 + i
        ws.cell(r, 1, fc).number_format = "#,##0"
        for j, n in enumerate(counts):
            f = (f"=IFERROR($A{r}/({n}*(1-{ref('変動費率')})),\"変動費率を入力\")")
            c = ws.cell(r, 2 + j, f)
            c.number_format = "#,##0"
            c.border = BOX

    _head(ws, 13, "実際の入力値での損益分岐", 4)
    items = [
        ("1件あたり粗利", f"={ref('初回単価')}*(1-{ref('変動費率')})"
                     f"+{ref('月額')}*{ref('継続月数')}*(1-{ref('変動費率')})", "円"),
        ("損益分岐に必要な月間受注件数", "=IFERROR(入力!$B$27/B14,\"\")", "件"),
        ("そのために必要な月間架電数", "=IFERROR(B15/入力!$B$25,\"\")", "本"),
        ("現在のキャパでの充足", "=IFERROR(入力!$B$24/B16,\"\")", "倍"),
        ("リストが尽きるまでの月数",
         f"=IFERROR({ref('リスト件数')}/入力!$B$24,\"\")", "ヶ月"),
    ]
    for i, (k, f, u) in enumerate(items):
        r = 14 + i
        ws.cell(r, 1, k).font = Font(bold=True)
        c = ws.cell(r, 2, f)
        c.border = BOX
        c.number_format = "#,##0.0" if u in ("倍", "ヶ月") else "#,##0"
        ws.cell(r, 3, u)
    ws.cell(17, 4, "1.0倍未満なら、今の人数では黒字化に届かない").font = \
        Font(size=10, color="808080")

    for col, w in zip("ABCDEFGH", (30, 14, 14, 14, 14, 14, 14, 40)):
        ws.column_dimensions[col].width = w


SCENARIOS = [
    ("現行の最終リスト", 1308, "民間×都内1事業所×独立系×全国1事業所"),
    ("＋除外を戻す", 1910, "併設・他県拠点・非営利を含める"),
    ("＋東京の多数事業所", 2893, "スイング側と合流。東京を全部使う"),
    ("＋全国に広げる", 36212, "同じ条件で全国。データは取得済み"),
]


def build_inventory(ws):
    ws["A1"] = "在庫シナリオ"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("流入経路がゼロのため、売上はリスト在庫に上限が決まります。"
                "どこまで広げるかが最大の論点です。")
    ws["A2"].font = Font(italic=True, color="808080")

    heads = ["広げ方", "件数", "枯渇まで(月)", "累計受注(目標値)",
             "累計売上", "内容"]
    for j, h in enumerate(heads):
        c = ws.cell(4, j + 1, h)
        c.font = WHITE_BOLD
        c.fill = HEAD
    for i, (lab, n, note) in enumerate(SCENARIOS):
        r = 5 + i
        ws.cell(r, 1, lab).font = Font(bold=True)
        ws.cell(r, 2, n).number_format = "#,##0"
        ws.cell(r, 3, f"=IFERROR(B{r}/入力!$B$24,\"\")").number_format = "#,##0.0"
        ws.cell(r, 4, f"=B{r}*入力!$B$25").number_format = "#,##0"
        ws.cell(r, 5, f"=D{r}*({ref('初回単価')}+{ref('月額')}*{ref('継続月数')})"
                ).number_format = "#,##0"
        ws.cell(r, 6, note).font = Font(size=10, color="808080")
        for j in range(1, 6):
            ws.cell(r, j).border = BOX
    for col, w in zip("ABCDEF", (24, 12, 14, 18, 16, 42)):
        ws.column_dimensions[col].width = w


ASSUMPTIONS = [
    ("■ この資料の性格", ""),
    ("", "金額はすべて未入力。埋めるまで損益はゼロのまま出ます。"
         "こちらで単価や人件費を推測して置くことはしていません。"),
    ("", ""),
    ("■ 検証済みの数字（確定）", ""),
    ("対象リスト 1,308件", "民間(営利法人)×都内1事業所×独立系×全国1事業所。"
                        "元データと照合済み"),
    ("東京都 全2,893件", "うち1法人1事業所が1,910件（法人番号で名寄せ）"),
    ("全国 36,212件", "全行が サービスの種類=居宅介護支援"),
    ("電話番号", "全件が元データと一致。欠損ゼロ"),
    ("", ""),
    ("■ 仮置きの数字（実測で置き換える）", ""),
    ("ファネル4率", "無効3% / アポ3% / 商談80% / 受注20%。"
                 "評価基準シートの目標値であり、業界の実績値ではない"),
    ("1人1日50件", "架電の所要時間を測っていないため仮。"
                "実測が出たら入力シートを更新する"),
    ("", ""),
    ("■ このモデルが示す構造的な論点", ""),
    ("リストは有限", "流入経路がゼロなので、売上の上限はリスト件数で決まる。"
                 "現行1,308件は3人×50件/日で約9営業日で尽きる"),
    ("売上が人数に比例する", "架電だけだと、売上を伸ばす手段が人を増やすことしかない。"
                        "利益率が構造的に上がらない"),
    ("枯渇後もコストは出る", "架電が止まっても人件費は消えない。"
                        "既定では枯渇後も固定費を計上する。"
                        "月次P&Lで赤字が続くのはその表現"),
    ("在庫を広げるか、経路を作るか", "全国に広げれば約12ヶ月分。"
                              "その間にSNS・Webの流入を立ち上げるかどうかが分岐点"),
    ("", ""),
    ("■ 使い方", ""),
    ("1", "入力シートの黄色いセルを埋める"),
    ("2", "損益分岐シートの上の表で、必要な単価が今の価格で届くか見る"),
    ("3", "月次P&Lで、何ヶ月目に累計黒字になるか確認する"),
    ("4", "在庫シナリオで、どこまでリストを広げるか決める"),
]


def build_assumptions(ws):
    ws["A1"] = "前提"
    ws["A1"].font = Font(bold=True, size=14)
    for i, (k, v) in enumerate(ASSUMPTIONS, start=3):
        a = ws.cell(i, 1, k or None)
        b = ws.cell(i, 2, v or None)
        if k.startswith("■"):
            a.font = WHITE_BOLD
            a.fill = HEAD
        else:
            a.font = Font(bold=True)
        b.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 82


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_assumptions(wb.create_sheet("前提"))
    build_input(wb.create_sheet("入力"))
    build_pl(wb.create_sheet("月次P&L"))
    build_breakeven(wb.create_sheet("損益分岐"))
    build_inventory(wb.create_sheet("在庫シナリオ"))
    wb.save(OUT)
    print(f"出力: {OUT}（シート {len(wb.sheetnames)}: {wb.sheetnames}）")
    print("金額セルは空欄です。入力シートの黄色いセルを埋めてください。")


if __name__ == "__main__":
    main()
