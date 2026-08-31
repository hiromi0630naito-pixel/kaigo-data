#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
厚労省「介護サービス情報公表システム」オープンデータ（居宅介護支援・全国版）から
東京都内に居宅介護支援事業所を1件だけ持つ法人を抽出し、架電リスト用Excelを生成する。

使い方:
    python3 scripts/extract_tokyo_1houjin_1kyotaku.py [CSVパス]

CSVパス省略時は既定URLからダウンロードを試みる。
"""

import os
import re
import subprocess
import sys
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CSV_URL = (
    "https://www.mhlw.go.jp/content/12300000/"
    "jigyosho_430_all_20260709180740.csv"
)
DEFAULT_CSV = "jigyosho_430_all.csv"
OUT_XLSX = "tokyo_1houjin_1kyotaku.xlsx"
OUT_CSV_LIST = "tokyo_minkan_1jigyosho.csv"     # 最終の架電リスト
OUT_CSV_EXCLUDED = "tokyo_jogai.csv"            # 除外した行と理由

# 事業所の基本情報（当初ご指定の並び順を維持し、代表者名・管理者名・エリアを追加）
BASE_COLS = [
    "No", "エリア", "都内1事業所", "全国1事業所",
    "法人種別", "都内事業所数", "全国事業所数", "事業所タイプ(推定)",
    "法人番号", "法人名", "代表者名", "管理者名", "法人所在地",
    "事業所名", "事業所番号", "事業所所在地",
    "電話番号", "電話番号(整形)", "電話種別", "FAX", "ホームページ",
    "利用可能曜日", "指定年月日",
]

# 元データに無くても残す列（架電時に手で埋めるため）
KEEP_IF_EMPTY = {"代表者名", "管理者名", "指定年月日"}

# 架電状況。集計でファネルに落とすため、互いに重ならない区分にしている。
ST_MITAKUSHU = "未着手"
ST_FUZAI = "不在・話中"          # 誰も出ない
ST_FUTSU = "番号違い・不通"      # 番号が無効。母数から外す
ST_BLOCK = "受付ブロック"        # 人は出たが担当に取り次がれない
ST_TALK = "担当者と会話"         # 担当と話せたがアポには至らず
ST_ORIKAESHI = "折り返し待ち"
ST_APO = "アポ獲得"
ST_KOTOWARI = "見込みなし・断り"
ST_TAISHOGAI = "対象外"          # サービス対象外。母数から外す

STATUS_ALL = [ST_MITAKUSHU, ST_FUZAI, ST_FUTSU, ST_BLOCK, ST_TALK,
              ST_ORIKAESHI, ST_APO, ST_KOTOWARI, ST_TAISHOGAI]
# 人につながった
ST_CONNECTED = [ST_BLOCK, ST_TALK, ST_ORIKAESHI, ST_APO, ST_KOTOWARI]
# 担当者まで到達した（受付で止まらなかった）
ST_REACHED = [ST_TALK, ST_ORIKAESHI, ST_APO, ST_KOTOWARI]
# 母数から除く（かけても評価できない相手）
ST_INVALID = [ST_FUTSU, ST_TAISHOGAI]

RESULT_JUCHU = "受注"
RESULT_ALL = ["未実施", "日程調整中", "実施済(検討中)", RESULT_JUCHU, "失注"]
RESULT_DONE = ["実施済(検討中)", RESULT_JUCHU, "失注"]   # 商談を実施した

# 架電時に記入する項目（空欄で出力）。メモ以外はすべてプルダウンにする。
CALL_COLS = [
    "記録者", "架電状況", "架電日", "架電回数", "対応者", "見込み度",
    "次回アクション", "次回架電予定日", "希望時間帯",
    "商談日", "商談結果", "通話メモ", "備考",
]

# 列名 -> プルダウンの選択肢
CHOICES = {
    "記録者": ["しょうちゃん", "さくや", "廣澤"],
    "架電状況": STATUS_ALL,
    "架電回数": ["1", "2", "3", "4", "5回以上"],
    "対応者": ["代表者", "管理者", "ケアマネ", "事務・受付", "その他", "不明"],
    "見込み度": ["A(即アポ)", "B(見込みあり)", "C(長期)", "D(見込み薄)"],
    "次回アクション": ["再架電", "資料送付", "メール送付", "訪問アポ", "対応不要"],
    "希望時間帯": ["午前", "13-15時", "15-17時", "17時以降", "指定なし"],
    "商談結果": RESULT_ALL,
}
# 日付として入力規則をかける列（プルダウンにはなじまないため）
DATE_COLS = ["架電日", "次回架電予定日", "商談日"]

# 幅を広めに取りたい自由入力欄
WIDE_COLS = {"通話メモ", "備考"}

HEADER_FILL_BASE = PatternFill("solid", fgColor="1F4E78")   # 濃紺: 基本情報
HEADER_FILL_CALL = PatternFill("solid", fgColor="C55A11")   # 橙  : 架電記入欄
HEADER_FONT = Font(color="FFFFFF", bold=True)


# --------------------------------------------------------------------------
# 1. 取得
# --------------------------------------------------------------------------
def ensure_csv(path):
    if os.path.exists(path):
        return path
    print(f"ダウンロード中: {CSV_URL}")
    subprocess.run(["curl", "-sSL", "--fail", "-o", path, CSV_URL], check=True)
    return path


# --------------------------------------------------------------------------
# 2. 読み込み（cp932 → utf-8-sig の順に試す）
# --------------------------------------------------------------------------
def find_header_row(path, encoding, max_scan=10):
    """先頭数行を覗いて、実際の見出し行のインデックスを返す。"""
    with open(path, "r", encoding=encoding, errors="strict") as f:
        for i, line in enumerate(f):
            if i >= max_scan:
                break
            if "事業所" in line and line.count(",") >= 5:
                return i
    return 0


def load_csv(path):
    last_err = None
    for enc in ("cp932", "utf-8-sig"):
        try:
            header_row = find_header_row(path, enc)
            df = pd.read_csv(
                path, encoding=enc, skiprows=header_row,
                dtype=str, low_memory=False, on_bad_lines="skip",
            )
            print(f"読み込み成功: encoding={enc}, 見出し行={header_row}, "
                  f"{len(df):,}行 x {len(df.columns)}列")
            return df
        except (UnicodeDecodeError, UnicodeError) as e:
            last_err = e
    raise RuntimeError(f"cp932 / utf-8-sig いずれでも読めませんでした: {last_err}")


# --------------------------------------------------------------------------
# 列名の解決（オープンデータは年度で列名が揺れるため部分一致で探す）
# --------------------------------------------------------------------------
KANA_EXCL = (r"かな", r"カナ", r"ｶﾅ", r"フリガナ", r"ふりがな")


def pick(cols, patterns, exclude=()):
    for pat in patterns:
        for c in cols:
            name = str(c)
            if re.search(pat, name) and not any(
                re.search(x, name) for x in exclude
            ):
                return c
    return None


def pick_all(cols, pattern, exclude=()):
    return [
        c for c in cols
        if re.search(pattern, str(c))
        and not any(re.search(x, str(c)) for x in exclude)
    ]


def join_parts(df, cols):
    """住所が都道府県/市区町村/番地/建物名に分割されている場合に連結する。

    方書（ビル名等）が住所側に既に含まれている形式のデータがあるため、
    既出の部分は連結せず読み飛ばす。
    """
    if not cols:
        return pd.Series([""] * len(df), index=df.index)
    acc = df[cols[0]].fillna("").astype(str).str.strip()
    for c in cols[1:]:
        part = df[c].fillna("").astype(str).str.strip()
        acc = pd.Series(
            [a if (not p or p in a) else a + p for a, p in zip(acc, part)],
            index=acc.index,
        )
    return acc.str.strip()


def resolve_columns(df):
    cols = list(df.columns)
    m = {}

    m["法人名"] = pick(
        cols, [r"法人.*名称", r"法人名", r"設置者.*名"], exclude=KANA_EXCL
    )
    # 代表者名: 「法人代表者」「代表者の氏名」「代表者名」等の揺れを吸収
    m["代表者名"] = pick(
        cols,
        [r"法人.*代表者.*氏名", r"代表者.*氏名", r"法人.*代表者名",
         r"代表者名", r"代表者"],
        exclude=KANA_EXCL + (r"職名", r"役職"),
    )
    # 管理者名: 架電時の取次先として有用なので拾えれば入れる
    m["管理者名"] = pick(
        cols,
        [r"管理者.*氏名", r"管理者名", r"事業所.*管理者"],
        exclude=KANA_EXCL + (r"職名", r"役職", r"経歴"),
    )
    m["法人番号"] = pick(cols, [r"^法人番号$", r"法人番号"], exclude=(r"コード",))
    m["利用可能曜日"] = pick(cols, [r"^利用可能曜日$", r"利用可能曜日"],
                        exclude=(r"特記",))
    m["事業所名"] = pick(cols, [r"事業所.*名称", r"事業所名"], exclude=KANA_EXCL)
    m["事業所番号"] = pick(cols, [r"事業所番号", r"事業所.*番号"])
    m["電話番号"] = pick(cols, [r"事業所.*電話", r"電話番号", r"電話"])
    m["FAX"] = pick(cols, [r"事業所.*FAX", r"FAX", r"ＦＡＸ", r"ファクシミリ"])
    m["ホームページ"] = pick(
        cols, [r"ホームページ", r"URL", r"ＵＲＬ", r"ウェブ", r"ＨＰ"]
    )
    m["指定年月日"] = pick(
        cols, [r"指定年月日", r"指定.*年月日", r"指定.*更新.*年月日"],
        exclude=(r"開始", r"休止", r"廃止", r"失効"),
    )

    houjin_addr = pick_all(cols, r"法人所在地", exclude=KANA_EXCL + (r"コード",))
    if not houjin_addr:
        c = pick(cols, [r"法人.*所在地", r"法人.*住所"], exclude=KANA_EXCL)
        houjin_addr = [c] if c else []
    m["_法人所在地_列"] = houjin_addr

    jigyo_addr = pick_all(cols, r"事業所所在地", exclude=KANA_EXCL + (r"コード",))
    if not jigyo_addr:
        # 「住所」＋「方書（ビル名等）」のように分かれている形式にも対応
        jigyo_addr = []
        c = pick(cols, [r"^住所$", r"事業所.*所在地", r"事業所.*住所",
                        r"所在地", r"住所"],
                 exclude=KANA_EXCL + (r"法人", r"コード"))
        if c:
            jigyo_addr.append(c)
        b = pick(cols, [r"方書", r"建物名", r"ビル名"], exclude=KANA_EXCL)
        if b:
            jigyo_addr.append(b)
    m["_事業所所在地_列"] = jigyo_addr

    m["_都道府県"] = pick(
        cols,
        [r"事業所所在地.*都道府県", r"^都道府県$", r"都道府県名", r"都道府県"],
        exclude=(r"コード", r"法人"),
    )
    # エリア（区市町村）専用列があれば最優先で使う
    m["_市区町村"] = pick(
        cols,
        [r"事業所所在地.*市区町村", r"事業所所在地.*市町村",
         r"^市区町村名$", r"^市区町村$", r"市区町村名", r"市町村名"],
        exclude=(r"コード", r"法人"),
    )
    return m


# --------------------------------------------------------------------------
# エリア（区市町村）の抽出
# --------------------------------------------------------------------------
def area_from_address(addr):
    """住所文字列から区市町村名を切り出す。専用列が無い場合のフォールバック。"""
    s = str(addr or "").strip()
    s = re.sub(r"^東京都", "", s)
    if "郡" in s:                      # 西多摩郡奥多摩町 -> 奥多摩町
        s = s.split("郡", 1)[1]
    # 「区」「市」を先に見る。非貪欲なので武蔵村山市を「武蔵村」で切らない
    for pat in (r"^(.+?[区市])", r"^(.+?[町村])"):
        mo = re.match(pat, s)
        if mo:
            return mo.group(1)
    return "不明"


def area_sort_key(name):
    """区 → 市 → 町村 → 不明 の順に並べる。"""
    if name.endswith("区"):
        return (0, name)
    if name.endswith("市"):
        return (1, name)
    if name.endswith(("町", "村")):
        return (2, name)
    return (3, name)


# --------------------------------------------------------------------------
# 正規化
# --------------------------------------------------------------------------
_FULL2HALF = str.maketrans(
    {chr(c): chr(c - 0xFEE0) for c in range(0xFF01, 0xFF5F)}
)


def normalize(s):
    """全角英数記号を半角化し、空白（全角含む）をすべて除去する。"""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = unicodedata.normalize("NFKC", str(s)).translate(_FULL2HALF)
    return re.sub(r"\s+", "", s).replace("　", "")


# --------------------------------------------------------------------------
# 架電リストからの除外条件（マーケティング上の理由）
# --------------------------------------------------------------------------
# 東京都の島嶼部。訪問できず架電効率も悪いため対象外にする。
ISLANDS = ["大島町", "利島村", "新島村", "神津島村", "三宅村",
           "御蔵島村", "八丈町", "青ヶ島村", "小笠原村"]


def phone_kind(p):
    """架電時の当たり方が変わるので電話番号の種類を分けておく。"""
    d = re.sub(r"\D", "", str(p or ""))
    if d.startswith(("090", "080", "070")):
        return "携帯"
    if d.startswith("050"):
        return "IP電話"
    if d.startswith("03"):
        return "固定(23区)"
    return "固定(多摩等)" if d else ""


def exclusion_reasons(d):
    """各行の除外理由を組み立てる。理由が空の行が架電対象。"""
    reasons = [[] for _ in range(len(d))]
    idx = {k: i for i, k in enumerate(d.index)}

    def mark(mask, label):
        for k in d.index[mask]:
            reasons[idx[k]].append(label)

    mark(d["事業所タイプ(推定)"] != "独立系", "併設事業所")
    mark(d["全国1事業所"] != "○", "他県にも拠点あり")
    mark(d["エリア"].isin(ISLANDS), "島嶼部")
    # 同じ番号への二重架電を防ぐため、2件目以降を落とす
    dg = d["電話番号(整形)"].map(lambda s: re.sub(r"\D", "", str(s)))
    mark(dg.duplicated(keep="first") & dg.ne(""), "電話番号が他社と重複")
    return pd.Series([" / ".join(r) for r in reasons], index=d.index)


# --------------------------------------------------------------------------
# 法人種別の判定
# --------------------------------------------------------------------------
# 法人名から法人格を判定する。上から順に評価するので、非営利・公的なものを
# 先に置き、営利法人は最後に判定する。
CORP_RULES = [
    ("公的機関", r"独立行政法人|国立(?!市)|日本赤十字社|^東京都|市役所|恩賜財団"),
    ("社協", r"社会福祉協議会"),
    ("社会福祉法人", r"社会福祉法人|シャカイフクシホウジン"),
    ("医療法人", r"医療法人|医療法社団"),
    ("公益法人", r"公益社団法人|公益財団法人"),
    ("学校・宗教", r"学校法人|宗教法人"),
    ("協同組合", r"協同組合"),
    ("NPO", r"特定非営利活動法人|特定非営利法人|NPO|ＮＰＯ"),
    ("一般社団・財団", r"一般社団法人|一般財団法人|社団法人|財団法人"),
    ("営利法人", r"株式会社|有限会社|合同会社|合資会社|合名会社"
               r"|\(株\)|（株）|\(有\)|（有）|㈱|㈲"),
]
# 民間企業として架電対象にする種別
MINKAN_TYPES = {"営利法人"}


def corp_type(name):
    """法人名から法人種別を判定する。法人格が名称に無い場合は不明とする。"""
    n = str(name or "")
    for label, pat in CORP_RULES:
        if re.search(pat, n):
            return label
    return "不明(法人格の記載なし)"


# --------------------------------------------------------------------------
# 事業所タイプの推定
# --------------------------------------------------------------------------
# このデータは全件が居宅介護支援の指定事業所だが、母体が病院や特養で、
# その一部門として居宅介護支援を持っている事業所も含まれる。
# 独立系のケアプラン事業所かどうかを事業所名から推定する（あくまで目安）。
FACILITY_PATTERNS = [
    ("医療機関併設", r"病院|医院|クリニック|診療所|外科|内科|医師会"),
    ("訪問看護併設", r"訪問看護"),
    ("施設併設", r"特別養護|養護老人|老人保健|老健|介護医療院|ケアハウス"
                r"|有料老人|サービス付き高齢者|ホーム(?!ヘルプ)|苑$|園$|荘$"),
    ("薬局併設", r"薬局|ドラッグ|薬品"),
    ("社協・包括", r"地域包括|社会福祉協議会|社協"),
]


def facility_type(name):
    """事業所名から母体施設を推定する。該当なしは独立系とみなす。"""
    n = str(name or "")
    for label, pat in FACILITY_PATTERNS:
        if re.search(pat, n):
            return label
    return "独立系"


# --------------------------------------------------------------------------
# 電話番号
# --------------------------------------------------------------------------
# 元データには全角数字・全角ハイフン・丸括弧・読点などの表記ゆれがある。
# 原文は改変せずそのまま残し、架電用に半角へ整形した列を別途用意する。
_SEP_RE = re.compile(r"[‐‑‒–—―−ー・.,、/()\[\]\s]+")


def phone_digits(s):
    """比較用。数字だけを取り出す。"""
    return re.sub(r"\D", "", unicodedata.normalize("NFKC", str(s or "")))


def normalize_phone(s):
    """半角数字とハイフンだけの表記に整える。数字は一切変えない。"""
    t = unicodedata.normalize("NFKC", str(s or "")).strip()
    t = _SEP_RE.sub("-", t)
    # 「直通03-...」のようなラベル文字を除去（数字は落とさない）
    t = re.sub(r"[^0-9\-]", "", t)
    t = re.sub(r"-{2,}", "-", t).strip("-")
    # 区切りが一切ない番号は、表計算ソフトが数値と解釈して先頭の0を落とす。
    # 桁の切り方が一意に決まる場合だけハイフンを補う（042系などは市外局番の
    # 桁数が一意でないため触らない）。
    if t.isdigit():
        if len(t) == 11 and t[:3] in ("050", "070", "080", "090"):
            t = f"{t[:3]}-{t[3:7]}-{t[7:]}"
        elif len(t) == 10 and t.startswith("03"):
            t = f"{t[:2]}-{t[2:6]}-{t[6:]}"
    # 整形で数字が変わっていないことを確認。変わるなら原文を返す（安全側）
    return t if phone_digits(t) == phone_digits(s) else str(s or "").strip()


def phone_audit(d):
    """架電前に人が目視すべき行を洗い出す。"""
    rows = []
    dg = d["電話番号"].map(phone_digits)
    dup_nums = set(dg.value_counts()[lambda v: v > 1].index) - {""}
    for i, r in d.iterrows():
        notes = []
        raw = str(r["電話番号"])
        if re.search(r"[^0-9\-]", raw):
            notes.append("表記ゆれ(全角・記号混在)")
        n = len(dg[i])
        if n == 0:
            notes.append("電話番号なし")
        elif n not in (10, 11):
            notes.append(f"桁数が異常({n}桁)")
        elif n == 11 and not dg[i].startswith(("050", "070", "080", "090")):
            notes.append("11桁だが携帯・IP以外")
        if dg[i] in dup_nums:
            notes.append("他の法人と同じ番号")
        if notes:
            rows.append({
                "No": r["No"], "エリア": r["エリア"], "法人名": r["法人名"],
                "事業所名": r["事業所名"], "電話番号(原文)": raw,
                "電話番号(整形)": r["電話番号(整形)"], "指摘": " / ".join(notes),
            })
    return pd.DataFrame(rows, columns=[
        "No", "エリア", "法人名", "事業所名",
        "電話番号(原文)", "電話番号(整形)", "指摘",
    ])


# --------------------------------------------------------------------------
# Excel 出力
# --------------------------------------------------------------------------
def display_width(v):
    """日本語を2、半角を1として概算幅を出す。"""
    return sum(
        2 if unicodedata.east_asian_width(ch) in ("F", "W", "A") else 1
        for ch in str(v)
    )


def safe_sheet_name(name, used):
    """Excelのシート名制約（31文字・記号禁止・重複不可）に合わせる。"""
    s = re.sub(r"[\[\]:*?/\\]", "", str(name))[:31] or "シート"
    base, i = s, 2
    while s in used:
        suffix = f"_{i}"
        s = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(s)
    return s


def style_sheet(ws, df, with_call_cols):
    """見出し装飾・1行目固定・オートフィルタ・列幅・入力規則をまとめて適用。"""
    ncol = len(df.columns)
    nrow = len(df)

    for i, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_CALL if col in CALL_COLS else HEADER_FILL_BASE
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if col in CALL_COLS:
            # 入力欄は中身が空なので見出し幅＋余白で決める
            width = max(display_width(col) + 4, 16)
            if col in WIDE_COLS:
                width = 30
        else:
            widths = [display_width(col)]
            widths += [display_width(v) for v in df[col].head(2000)]
            width = min(max(widths) + 2, 60)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"                                  # 1行目固定
    if nrow:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{nrow + 1}"

    if not (with_call_cols and nrow):
        return

    # メモ以外の入力欄はプルダウン、日付欄は日付の入力規則にする
    cols = list(df.columns)
    for col_name, choices in CHOICES.items():
        if col_name not in cols:
            continue
        letter = get_column_letter(cols.index(col_name) + 1)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(choices) + '"',
            allow_blank=True,
            showDropDown=False,   # False で「ドロップダウンを表示する」がON
            promptTitle=col_name,
            prompt="リストから選択してください",
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{nrow + 1}")

    for col_name in DATE_COLS:
        if col_name not in cols:
            continue
        letter = get_column_letter(cols.index(col_name) + 1)
        dv = DataValidation(
            type="date", operator="between",
            formula1="DATE(2020,1,1)", formula2="DATE(2035,12,31)",
            allow_blank=True,
            promptTitle=col_name, prompt="日付を入力（例 2026/9/1）",
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{nrow + 1}")
        for r in range(2, nrow + 2):
            ws.cell(r, cols.index(col_name) + 1).number_format = "yyyy/m/d"


def write_excel(path, full, target, strict, call_list, excluded,
                summary, areas, audit):
    used = set()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        sheets = [
            ("架電リスト(最終)", call_list, True),
            ("除外リスト", excluded, False),
            ("1法人1事業所(都内)", target, True),
            ("1法人1事業所(全国)", strict, True),
            ("架電リスト(全件)", full, True),
            ("エリア別集計", summary, False),
            ("電話番号_要確認", audit, False),
        ]
        # 区市町村ごとのシートは全件を分割（1法人1事業所は列で絞り込める）
        for area in areas:
            sheets.append((area, full[full["エリア"] == area], True))

        for name, df, with_call in sheets:
            sheet = safe_sheet_name(name, used)
            df.to_excel(writer, sheet_name=sheet, index=False)
            style_sheet(writer.sheets[sheet], df, with_call)

        # 評価用のシート。架電リストの列順が確定した後に作る必要がある。
        wb = writer.book
        write_dashboard(wb.create_sheet("評価ダッシュボード"),
                        list(call_list.columns),
                        sorted(call_list["エリア"].unique(), key=area_sort_key))
        write_criteria(wb.create_sheet("評価基準"))
        # 使う順に並べ替える。架電リスト → ダッシュボード → 評価基準 → その他
        order = [SHEET_LIST, "評価ダッシュボード", "評価基準"]
        by_name = {w.title: w for w in wb._sheets}
        wb._sheets = ([by_name[n] for n in order if n in by_name]
                      + [w for w in wb._sheets if w.title not in order])


# --------------------------------------------------------------------------
def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_CSV
    df = load_csv(ensure_csv(csv_path))
    m = resolve_columns(df)

    missing = [k for k in ("法人名", "事業所名") if not m.get(k)]
    if missing:
        raise RuntimeError(
            f"必須列が見つかりません: {missing}\n列一覧: {list(df.columns)}"
        )

    def col(key):
        c = m.get(key)
        return (df[c].fillna("").astype(str).str.strip()
                if c else pd.Series([""] * len(df), index=df.index))

    out = pd.DataFrame(index=df.index)
    out["法人番号"] = col("法人番号")
    out["法人名"] = col("法人名")
    out["代表者名"] = col("代表者名")
    out["管理者名"] = col("管理者名")
    out["法人所在地"] = join_parts(df, m["_法人所在地_列"])
    out["事業所名"] = col("事業所名")
    out["事業所番号"] = col("事業所番号")
    out["事業所所在地"] = join_parts(df, m["_事業所所在地_列"])
    for k in ("電話番号", "FAX", "ホームページ", "利用可能曜日", "指定年月日"):
        out[k] = col(k)
    out["電話番号(整形)"] = out["電話番号"].map(normalize_phone)
    out["事業所タイプ(推定)"] = out["事業所名"].map(facility_type)
    out["電話種別"] = out["電話番号(整形)"].map(phone_kind)
    # 法人格は名称に書かれていないことがある。同じ法人番号の行が全国に
    # あれば、そちらの名称に法人格が書かれている場合があるので突き合わせる。
    # 例: 東京の行が「杉の子」でも、他県の行が「社会福祉法人　杉の子」。
    names_by_no = (
        out.loc[out["法人番号"] != "", ["法人番号", "法人名"]]
        .groupby("法人番号")["法人名"].apply(lambda v: " ".join(sorted(set(v))))
    )
    merged_name = out["法人番号"].map(names_by_no).fillna(out["法人名"])
    out["法人種別"] = merged_name.where(
        out["法人番号"] != "", out["法人名"]).map(corp_type)

    # 全国での事業所数。東京都で1事業所でも他県に拠点を持つ法人を見分ける。
    nat = out.loc[out["法人番号"] != "", "法人番号"].value_counts()
    out["全国事業所数"] = out["法人番号"].map(nat).fillna(0).astype(int)

    # 整形で数字が変化していないか全件検証（1件でもあれば異常として止める）
    changed = (out["電話番号"].map(phone_digits)
               != out["電話番号(整形)"].map(phone_digits))
    if changed.any():
        raise RuntimeError(f"電話番号の整形で数字が変化しました: {int(changed.sum())}件")
    print(f"  電話番号の整形: {int((out['電話番号'] != out['電話番号(整形)']).sum()):,}件を"
          f"半角化（数字の変化なしを全件検証済み）")

    # 元データに無かった列を報告
    for label in ("法人番号", "代表者名", "管理者名", "電話番号", "FAX",
                  "ホームページ", "利用可能曜日", "指定年月日"):
        if not m.get(label):
            print(f"  ※ 注意: 「{label}」に対応する列が元データに見つからず、"
                  f"空欄で出力します")
        else:
            print(f"  {label:<6} <- 元CSV列「{m[label]}」")

    # ---- 3. 東京都に絞り込み -------------------------------------------
    if m["_都道府県"]:
        pref = df[m["_都道府県"]].fillna("").astype(str).str.strip()
        mask = pref.str.startswith("東京都") | (pref == "東京")
        how = f'都道府県列「{m["_都道府県"]}」'
    else:
        mask = out["事業所所在地"].str.startswith("東京都")
        how = "事業所所在地の先頭一致"
    tokyo = out[mask].copy()
    print(f"東京都の絞り込み方法: {how}")

    # ---- エリア（区市町村） --------------------------------------------
    if m["_市区町村"]:
        tokyo["エリア"] = (
            df.loc[tokyo.index, m["_市区町村"]].fillna("").astype(str).str.strip()
        )
        tokyo.loc[tokyo["エリア"] == "", "エリア"] = (
            tokyo.loc[tokyo["エリア"] == "", "事業所所在地"].map(area_from_address)
        )
        # 「西多摩郡奥多摩町」→「奥多摩町」に揃える（住所切り出し側と統一）
        tokyo["エリア"] = tokyo["エリア"].str.replace(r"^.*郡", "", regex=True)
        print(f'エリアの取得元: 市区町村列「{m["_市区町村"]}」')
    else:
        tokyo["エリア"] = tokyo["事業所所在地"].map(area_from_address)
        print("エリアの取得元: 事業所所在地から切り出し")

    # ---- 4. 法人名＋法人所在地でグルーピング ---------------------------
    # 法人番号があればそれが最も正確な同一法人の判定。
    # 空欄の行だけ、法人名＋法人所在地の正規化キーで補う。
    name_key = (
        "名:" + tokyo["法人名"].map(normalize)
        + "|" + tokyo["法人所在地"].map(normalize)
    )
    hojin_no = tokyo["法人番号"].map(normalize)
    key = hojin_no.where(hojin_no == "", "番:" + hojin_no)
    key = key.mask(hojin_no == "", name_key)
    n_no = int((hojin_no != "").sum())
    print(f"法人の名寄せキー: 法人番号 {n_no:,}件 / "
          f"法人名＋法人所在地 {len(tokyo) - n_no:,}件")
    counts = key.map(key.value_counts())
    tokyo["都内事業所数"] = counts.astype(int)
    tokyo["都内1事業所"] = counts.eq(1).map({True: "○", False: ""})
    tokyo["全国1事業所"] = tokyo["全国事業所数"].eq(1).map({True: "○", False: ""})

    # ---- 並べ替え・列整形 ----------------------------------------------
    def finalize(d, with_call):
        d = d.copy()
        d["_k"] = d["エリア"].map(area_sort_key)
        d = d.sort_values(["_k", "法人名", "事業所名"]).drop(columns="_k")
        d = d.reset_index(drop=True)
        d.insert(0, "No", range(1, len(d) + 1))
        cols = [
            c for c in BASE_COLS
            if c in KEEP_IF_EMPTY or c in ("No", "エリア")
            or d[c].astype(str).str.strip().ne("").any()
        ]
        if with_call:
            for c in CALL_COLS:
                d[c] = ""
            cols += CALL_COLS
        return d[cols]

    # 全件を通し番号つきで確定させ、1法人1事業所はその部分集合として切り出す。
    # こうすることで No がどのシートでも同じ行を指す。
    full = finalize(tokyo, True)
    target = full[full["都内1事業所"] == "○"].copy()
    strict = target[target["全国1事業所"] == "○"].copy()

    areas = sorted(full["エリア"].unique(), key=area_sort_key)
    summary = (
        pd.DataFrame({"エリア": areas})
        .assign(
            全件数=lambda d: d["エリア"].map(
                full["エリア"].value_counts()).fillna(0).astype(int),
            都内1事業所=lambda d: d["エリア"].map(
                target["エリア"].value_counts()).fillna(0).astype(int),
            全国でも1事業所=lambda d: d["エリア"].map(
                strict["エリア"].value_counts()).fillna(0).astype(int),
            うち独立系=lambda d: d["エリア"].map(
                strict[strict["事業所タイプ(推定)"] == "独立系"]["エリア"]
                .value_counts()).fillna(0).astype(int),
        )
    )

    # ---- 5. Excel 出力 --------------------------------------------------
    audit = phone_audit(full)
    # 民間（営利法人）のみを切り出す。法人格が名称に無い行は判定できないため
    # 別枠にして、人が見て判断できるようにする。
    minkan = target[target["法人種別"].isin(MINKAN_TYPES)].copy()
    minkan["除外理由"] = exclusion_reasons(minkan)
    call_list = minkan[minkan["除外理由"] == ""].drop(columns="除外理由")

    # 除外した行は理由つきで残す。法人格が名称に無く民間と判定できなかった
    # 行もここにまとめる。
    dropped = minkan[minkan["除外理由"] != ""].copy()
    other = target[~target["法人種別"].isin(MINKAN_TYPES)].copy()
    other["除外理由"] = "民間(営利法人)ではない: " + other["法人種別"]
    excluded = pd.concat([dropped, other], ignore_index=True)
    cols = [c for c in minkan.columns if c != "除外理由"]
    excluded = excluded[["除外理由"] + cols]

    print("\n  架電リストからの除外内訳:")
    for label, n in (
        minkan.loc[minkan["除外理由"] != "", "除外理由"]
        .str.split(" / ").explode().value_counts().items()
    ):
        print(f"    {label:<20}: {n:>4} 件")

    write_excel(OUT_XLSX, full, target, strict, call_list, excluded,
                summary, areas, audit)

    # CSVは他ツールにそのまま取り込めるように出す。
    # Excelでそのまま開けるよう utf-8-sig（BOM付き）にする。
    for path, d in ((OUT_CSV_LIST, call_list), (OUT_CSV_EXCLUDED, excluded)):
        d.to_csv(path, index=False, encoding="utf-8-sig")
        print(f"  CSV出力: {path} ({len(d):,}件)")

    # ---- 6. 報告 --------------------------------------------------------
    print()
    print(f"東京都の居宅介護支援 総事業所数: {len(full):,} 件")
    print(f"1法人1事業所の件数(都内基準)   : {len(target):,} 件")
    print(f"  うち全国でも1事業所          : {len(strict):,} 件")
    print(f"  うち独立系(併設でない)と推定  : "
          f"{int((strict['事業所タイプ(推定)'] == '独立系').sum()):,} 件")
    print(f"  ※都内1事業所だが他県に拠点あり: {len(target) - len(strict):,} 件")
    print("\n  法人種別の内訳(都内1事業所):")
    for k, v in target["法人種別"].value_counts().items():
        mark = " ← 架電対象" if k in MINKAN_TYPES else ""
        print(f"    {k:<20}: {v:>5} 件{mark}")
    print(f"エリア数                      : {len(areas)} 区市町村")
    print(f"電話番号 要確認                : {len(audit):,} 件"
          f"（数字自体の誤りではなく表記ゆれ等）")
    print(f"出力: {OUT_XLSX}")



# --------------------------------------------------------------------------
# 評価ダッシュボード
# --------------------------------------------------------------------------
# 架電リストが埋まると自動集計される数式を置く。Excelでも Googleスプレッド
# シートでも動くよう COUNTIF / COUNTIFS / IFERROR だけで組む。
SHEET_LIST = "架電リスト(最終)"
LAST_ROW = 5000

# KPIと初期の判定ライン。実測が貯まったら置き換える前提の仮置き。
KPI_TARGETS = [
    ("接続率", 0.50, "有効架電のうち人につながった割合"),
    ("担当者到達率", 0.60, "接続のうち受付で止まらず担当と話せた割合"),
    ("アポ率(有効架電比)", 0.03, "架電1本あたりのアポ獲得効率"),
    ("アポ率(接続比)", 0.06, "つながった相手からアポを取れた割合"),
    ("商談実施率", 0.80, "アポのうち実際に商談できた割合"),
    ("受注率(商談比)", 0.20, "商談のうち受注に至った割合"),
    ("無効リード率(上限)", 0.03, "番号違い・対象外の割合。低いほどよい"),
]
MIN_SAMPLE = 100      # セグメントを評価してよい最低架電数
STRONG_RATIO = 1.3    # 全体平均のこの倍以上なら強化
WEAK_RATIO = 0.6      # 全体平均のこの倍以下なら縮小検討


def _rng(cols, name):
    """列名から架電リストの絶対参照レンジを作る。"""
    L = get_column_letter(cols.index(name) + 1)
    return "'" + SHEET_LIST + "'!$" + L + "$2:$" + L + "$" + str(LAST_ROW)


def _any(rng, values):
    """いずれかの値に一致する件数。COUNTIFの和で組む。"""
    return "+".join('COUNTIF(' + rng + ',"' + v + '")' for v in values)


def _seg_block(label, rng, values, r_st, avg_cell, row0):
    """セグメント別ブロックの行を返す。row0 はこのブロックの先頭行番号。"""
    out = [["■ " + label + "別", None, None, None, None, None],
           ["区分", "架電済", "接続", "アポ", "アポ率", "判定"]]
    for i, v in enumerate(values):
        r = row0 + 2 + i
        conn = "+".join('COUNTIFS(' + rng + ',"' + v + '",' + r_st + ',"' + s + '")'
                        for s in ST_CONNECTED)
        out.append([
            v,
            '=COUNTIFS(' + rng + ',"' + v + '",' + r_st + ',"<>' + ST_MITAKUSHU + '")',
            "=" + conn,
            '=COUNTIFS(' + rng + ',"' + v + '",' + r_st + ',"' + ST_APO + '")',
            '=IFERROR(D' + str(r) + '/B' + str(r) + ',"")',
            '=IF(B' + str(r) + '<' + str(MIN_SAMPLE) + ',"サンプル不足",'
            'IF(E' + str(r) + '="","-",'
            'IF(E' + str(r) + '>=' + avg_cell + '*' + str(STRONG_RATIO) + ',"強化",'
            'IF(E' + str(r) + '<=' + avg_cell + '*' + str(WEAK_RATIO) + ',"縮小検討",'
            '"継続"))))',
        ])
    out.append([None] * 6)
    return out


def build_dashboard_rows(cols, areas):
    """ダッシュボードの全行と、パーセント書式にする行番号を返す。"""
    st = _rng(cols, "架電状況")
    rows, pct = [], []

    def add(*vals):
        rows.append(list(vals) + [None] * (6 - len(vals)))

    add("■ 全体ファネル")
    add("段階", "件数", "通過率", "説明")
    add("対象件数", "=COUNTA(" + _rng(cols, "エリア") + ")", None, "リストの総数")
    add("架電済", "=COUNTA(" + st + ')-COUNTIF(' + st + ',"' + ST_MITAKUSHU + '")',
        '=IFERROR(B4/B3,"")', "未着手を除く")
    add("有効架電", "=B4-(" + _any(st, ST_INVALID) + ")",
        '=IFERROR(B5/B4,"")', "番号違い・対象外を除いた母数")
    add("接続(人が出た)", "=" + _any(st, ST_CONNECTED),
        '=IFERROR(B6/B5,"")', "← 接続率")
    add("担当者到達", "=" + _any(st, ST_REACHED),
        '=IFERROR(B7/B6,"")', "受付で止まらなかった")
    add("決裁者と会話",
        '=COUNTIF(' + _rng(cols, "対応者") + ',"代表者")+COUNTIF('
        + _rng(cols, "対応者") + ',"管理者")',
        '=IFERROR(B8/B7,"")', "対応者が代表者・管理者")
    add("アポ獲得", '=COUNTIF(' + st + ',"' + ST_APO + '")',
        '=IFERROR(B9/B8,"")', "← ゴール")
    add("商談実施", "=" + _any(_rng(cols, "商談結果"), RESULT_DONE),
        '=IFERROR(B10/B9,"")')
    add("受注", '=COUNTIF(' + _rng(cols, "商談結果") + ',"' + RESULT_JUCHU + '")',
        '=IFERROR(B11/B10,"")')
    pct += [4, 5, 6, 7, 8, 9, 10, 11]
    add()

    add("■ KPIと判定ライン")
    add("KPI", "実測", "目標(仮)", "判定", "説明")
    kf = ['=IFERROR(B6/B5,"")', '=IFERROR(B7/B6,"")', '=IFERROR(B9/B5,"")',
          '=IFERROR(B9/B6,"")', '=IFERROR(B10/B9,"")', '=IFERROR(B11/B10,"")',
          '=IFERROR((' + _any(st, ST_INVALID) + ')/B4,"")']
    for i, ((name, target, note), f) in enumerate(zip(KPI_TARGETS, kf)):
        r = 15 + i
        if "上限" in name:
            j = ('=IF(B' + str(r) + '="","未計測",IF(B' + str(r)
                 + '<=C' + str(r) + ',"OK","要改善"))')
        else:
            j = ('=IF(B' + str(r) + '="","未計測",IF(B' + str(r)
                 + '>=C' + str(r) + ',"OK","未達"))')
        add(name, f, target, j, note)
        pct.append(r)
    APO = "B17"          # アポ率(有効架電比) のセル。セグメント判定の基準
    add()
    add("※ 目標値は初期の仮置き。300件ほど架電したら実測で置き換えること。")
    add()

    # --- セグメント別 ---
    segs = [
        ("電話種別", _rng(cols, "電話種別"),
         ["固定(23区)", "固定(多摩等)", "携帯", "IP電話"]),
        ("記録者", _rng(cols, "記録者"), CHOICES["記録者"]),
        ("見込み度", _rng(cols, "見込み度"), CHOICES["見込み度"]),
        ("対応者", _rng(cols, "対応者"), CHOICES["対応者"]),
        ("希望時間帯", _rng(cols, "希望時間帯"), CHOICES["希望時間帯"]),
        ("エリア", _rng(cols, "エリア"), areas),
    ]
    for label, rng, vals in segs:
        row0 = len(rows) + 1
        block = _seg_block(label, rng, vals, st, APO, row0)
        for b in block:
            rows.append(b)
        pct += [row0 + 2 + i for i in range(len(vals))]
    return rows, pct


def write_dashboard(ws, cols, areas):
    """評価ダッシュボードのシートを組み立てる。"""
    rows, pct = build_dashboard_rows(cols, areas)
    for r, row in enumerate(rows, start=1):
        for cix, v in enumerate(row, start=1):
            if v is None:
                continue
            cell = ws.cell(r, cix, v)
            if isinstance(v, str) and v.startswith("■"):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = HEADER_FILL_BASE
            elif isinstance(v, str) and v.startswith("※"):
                cell.font = Font(italic=True, color="808080")
    for r in pct:
        for cix in (3, 5):
            ws.cell(r, cix).number_format = "0.0%"
        if r <= 21:
            ws.cell(r, 2).number_format = "0.0%"
            ws.cell(r, 3).number_format = "0.0%"
    for r in range(4, 12):
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "0.0%"
    widths = [26, 14, 12, 34, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


CRITERIA_DOC = [
    ("■ この評価基準の使い方", ""),
    ("", "架電を始める前に判定ラインを決めておき、"
         "貯まった実測で毎週見直す。"),
    ("", "数値は「評価ダッシュボード」シートが自動集計する。"),
    ("", ""),
    ("■ 1. 何を測るか（ファネル）", ""),
    ("対象件数", "リストの総数。分母の起点"),
    ("架電済", "1回でもかけた件数。未着手を除く"),
    ("有効架電", "架電済から番号違い・対象外を除く。"
                "リストの質が悪いと母数が減る"),
    ("接続", "人が出た件数。時間帯とリストの鮮度で決まる"),
    ("担当者到達", "受付で止まらず担当と話せた件数。"
                  "トークの入口の強さ"),
    ("決裁者と会話", "対応者が代表者・管理者。"
                    "1法人1事業所を狙う理由がここに出る"),
    ("アポ獲得", "ゴール。ただし数だけ見ると質が落ちるので受注まで追う"),
    ("商談実施", "アポのうち実際に会えた件数。ドタキャン率が見える"),
    ("受注", "最終成果"),
    ("", ""),
    ("■ 2. 判定ライン（初期は仮置き）", ""),
    ("接続率 50%", "下回るなら架電時間帯を見直す。"
                  "土日祝営業の事業所は平日日中が捕まりにくい"),
    ("担当者到達率 60%", "下回るなら受付突破のトークが弱い。"
                        "名乗りと用件の一言目を変える"),
    ("アポ率(有効架電比) 3%", "全体の効率指標。"
                            "1アポあたり約33本が目安になる"),
    ("アポ率(接続比) 6%", "話せた相手を口説けているか。"
                        "接続率と分けて見ることでボトルネックが特定できる"),
    ("商談実施率 80%", "下回るならアポの質が低い。"
                      "その場の勢いだけで取れていないか"),
    ("受注率(商談比) 20%", "商材とターゲットの適合。"
                          "低いならリスト条件そのものを疑う"),
    ("無効リード率 3%以下", "上回るならリストの鮮度に問題。"
                          "元データの更新を検討"),
    ("", ""),
    ("■ 3. いつ判断してよいか", ""),
    ("最低100件", "セグメントごとに架電100件を超えるまでは判定しない。"),
    ("", "アポ率が数%の世界なので、"
         "50件程度の差はほぼ偶然で説明がついてしまう。"),
    ("全体300件", "全体で300件を超えたら、"
                 "上の目標値を自分たちの実測値に置き換える。"),
    ("", ""),
    ("■ 4. セグメントの打ち切り基準", ""),
    ("強化", "そのセグメントのアポ率が全体平均の1.3倍以上。"
            "残りリストの配分を増やす"),
    ("継続", "全体平均の0.6〜1.3倍。現状維持"),
    ("縮小検討", "全体平均の0.6倍以下。"
                "件数が100を超えていれば配分を減らす"),
    ("サンプル不足", "架電100件未満。判断しない"),
    ("", ""),
    ("■ 5. 見るべきセグメント", ""),
    ("電話種別", "携帯・IP電話は代表直通の可能性が高い。"
                "固定(23区)との差が出るかが最初の検証点"),
    ("記録者", "3人の差はトークの差。"
              "差が出たら勝ちパターンを共有して揃える"),
    ("見込み度", "A評価がどれだけ受注に化けたか。"
                "化けないなら見込み度の付け方がずれている"),
    ("対応者", "誰に当たった時に決まるか。"
              "受付で粘るべきか切るべきかの判断材料"),
    ("希望時間帯", "折り返しがどの時間帯に集中するか。"
                  "翌週の架電計画に反映する"),
    ("エリア", "地域差。訪問効率とセットで見る"),
    ("", ""),
    ("■ 6. 週次で見る順番", ""),
    ("① ファネルのどこが細いか", "接続→担当到達→アポの通過率を上から見る"),
    ("② 一番細い段階だけ改善する", "同時に複数を変えると"
                                  "何が効いたか分からなくなる"),
    ("③ セグメント判定を更新", "強化・縮小検討の配分を翌週に反映"),
    ("④ 目標値の見直し", "実測が目標を安定して超えたら目標を上げる"),
    ("", ""),
    ("■ 7. 注意", ""),
    ("アポ数だけ追わない", "アポ率は上がったのに受注率が落ちる、"
                          "が最もありがちな失敗"),
    ("1件目の架電で判断しない", "不在が続くのは普通。"
                              "架電回数と接触率の関係も見る"),
    ("除外リストを定期的に見る", "併設や他県拠点を切った判断が"
                                "正しかったか、後から検証できる"),
]


def write_criteria(ws):
    """評価基準の説明シートを書く。"""
    ws.cell(1, 1, "評価基準").font = Font(bold=True, size=14)
    for i, (k, v) in enumerate(CRITERIA_DOC, start=3):
        a, b = ws.cell(i, 1, k or None), ws.cell(i, 2, v or None)
        if k.startswith("■"):
            a.font = Font(bold=True, color="FFFFFF")
            a.fill = HEADER_FILL_BASE
        else:
            a.font = Font(bold=True)
        b.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 78


if __name__ == "__main__":
    main()
